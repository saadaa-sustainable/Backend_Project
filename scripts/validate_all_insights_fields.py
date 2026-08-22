"""Test EVERY candidate Insights field, AT EVERY LEVEL (account / campaign
/ adset / ad), individually against the real Meta API and produce a
per-field-per-level pass/fail report -- not batch-level like
dump_test_table.py's field-batching, but one entry per (field, level):

    {"impressions": {"account_level": "passed", "campaign_level": "passed",
                      "adset_level": "passed", "ad_level": "passed"},
     "quality_ranking": {"account_level": "error 100: ...", ...,
                          "ad_level": "passed"}, ...}

Field universe = every field currently in the registry
(app/core/meta_registry.py's INSIGHTS_FIELD_GROUPS) UNION every field
listed in exports/meta_ads_insights_full_fields_reference.xlsx (Meta's
documented "Fields" section, ~289 names) -- i.e. every field this project
has ever considered, whether currently included or not.

Each (field, level) pair is requested completely on its own (limit=1,
date_preset=yesterday) so a failure can never be blamed on a neighboring
field or level -- slower than dump_test_table.py's batching (one API call
per field per level, bounded concurrency: 4x the calls of the field-only
version) but gives an unambiguous verdict per field per level, matching
what a shareable audit log needs. Also empirically discovers/confirms
level restrictions (see LEVEL_RESTRICTED_METRICS in meta_registry.py)
rather than relying on the registry's own documentation of them.

Output (never touches the registry itself):
    logs/insights_field_validation.json   -- {field: {level: "passed"|"error: ..."}}
    logs/insights_field_validation.csv    -- one row per (field, level), full detail
    logs/insights_field_validation.jsonl  -- one JSON object per (field, level)

Usage:
    python3 scripts/validate_all_insights_fields.py                # account 1, yesterday, all 4 levels
    python3 scripts/validate_all_insights_fields.py --account 2
    python3 scripts/validate_all_insights_fields.py --levels ad,campaign
    python3 scripts/validate_all_insights_fields.py --concurrency 10
"""

from __future__ import annotations

import argparse
import ast
import asyncio
import csv
import json
import os
import sys
import time
from pathlib import Path
from typing import Any

try:
    import httpx
except ImportError:
    print("Missing dependency: pip install httpx", file=sys.stderr)
    raise SystemExit(1)

try:
    from openpyxl import load_workbook
except ImportError:
    print("Missing dependency: pip install openpyxl", file=sys.stderr)
    raise SystemExit(1)

try:
    from dotenv import load_dotenv
except ImportError:
    load_dotenv = None

sys.path.insert(0, str(Path(__file__).resolve().parent))
from dump_test_table import (  # noqa: E402
    REPO_ROOT,
    META_REGISTRY_PATH,
    _discover_accounts,
    _get_with_retry,
    _parse_meta_error,
    AccountConfig,
)

FIELDS_XLSX_PATH = REPO_ROOT / "exports" / "meta_ads_insights_full_fields_reference.xlsx"
OUT_JSON_PATH = REPO_ROOT / "logs" / "insights_field_validation.json"
OUT_CSV_PATH = REPO_ROOT / "logs" / "insights_field_validation.csv"
OUT_JSONL_PATH = REPO_ROOT / "logs" / "insights_field_validation.jsonl"
CSV_COLUMNS = [
    "field", "level", "status", "in_current_registry",
    "error_code", "error_subcode", "error_type", "error_message",
]
#: Matches app.core.meta_registry.InsightsLevel's values.
ALL_LEVELS = ["account", "campaign", "adset", "ad"]


def _load_registry_field_universe(registry_path: Path) -> set[str]:
    """Every field name anywhere in INSIGHTS_FIELD_GROUPS, including
    isolation-only groups (skadnetwork, configurable_reach_and_attribution)
    -- this is "everything the registry currently knows about", not just
    the default fetch set."""
    tree = ast.parse(registry_path.read_text())
    for node in tree.body:
        target = None
        if isinstance(node, ast.AnnAssign) and isinstance(node.target, ast.Name):
            target = node.target.id
        elif isinstance(node, ast.Assign) and isinstance(node.targets[0], ast.Name):
            target = node.targets[0].id
        if target == "INSIGHTS_FIELD_GROUPS":
            groups: dict[str, list[str]] = ast.literal_eval(node.value)
            return {f for fields in groups.values() for f in fields}
    raise RuntimeError(f"INSIGHTS_FIELD_GROUPS not found in {registry_path}")


def _load_doc_field_universe(xlsx_path: Path) -> set[str]:
    wb = load_workbook(xlsx_path, read_only=True)
    ws = wb["Ads Insights - All Fields"]
    names = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        name = row[1]  # column B = Field Name
        if name:
            names.add(str(name))
    return names


async def _test_one_field_level(
    client: httpx.AsyncClient,
    base_url: str,
    account: AccountConfig,
    access_token: str,
    field_name: str,
    level: str,
    date_preset: str,
    semaphore: asyncio.Semaphore,
) -> dict[str, Any]:
    params = {
        "access_token": access_token,
        "level": level,
        "fields": field_name,
        "date_preset": date_preset,
        "limit": 1,
    }
    async with semaphore:
        try:
            await _get_with_retry(client, f"{base_url}/act_{account.account_id}/insights", params)
            return {"field": field_name, "level": level, "status": "passed"}
        except RuntimeError as exc:
            parsed = _parse_meta_error(str(exc))
            return {
                "field": field_name,
                "level": level,
                "status": "error",
                "error_code": parsed.get("error_code", ""),
                "error_subcode": parsed.get("error_subcode", ""),
                "error_type": parsed.get("error_type", ""),
                "error_message": parsed.get("error_message") or str(exc)[:300],
            }


async def _run(
    fields: list[str],
    levels: list[str],
    *,
    base_url: str,
    account: AccountConfig,
    access_token: str,
    date_preset: str,
    concurrency: int,
) -> list[dict[str, Any]]:
    semaphore = asyncio.Semaphore(concurrency)
    limits = httpx.Limits(max_connections=concurrency + 5, max_keepalive_connections=concurrency)
    async with httpx.AsyncClient(limits=limits) as client:
        tasks = [
            _test_one_field_level(client, base_url, account, access_token, f, level, date_preset, semaphore)
            for f in fields
            for level in levels
        ]
        results = []
        done = 0
        for coro in asyncio.as_completed(tasks):
            result = await coro
            done += 1
            if done % 100 == 0 or done == len(tasks):
                print(f"  {done}/{len(tasks)} (field, level) pairs tested...")
            results.append(result)
        return results


def main() -> int:
    sys.stdout.reconfigure(line_buffering=True)

    parser = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    parser.add_argument("--account", default=None, help="Account key to test against (default: first configured).")
    parser.add_argument("--date-preset", default="yesterday", help="Meta date_preset (default: yesterday).")
    parser.add_argument(
        "--levels", default=",".join(ALL_LEVELS),
        help=f"Comma-separated Insights levels to test each field at (default: all -- {','.join(ALL_LEVELS)}).",
    )
    parser.add_argument("--concurrency", type=int, default=8, help="Concurrent (field, level) requests (default 8).")
    args = parser.parse_args()
    levels = [lvl.strip() for lvl in args.levels.split(",") if lvl.strip()]
    unknown_levels = set(levels) - set(ALL_LEVELS)
    if unknown_levels:
        print(f"Unknown level(s): {sorted(unknown_levels)}. Valid: {ALL_LEVELS}", file=sys.stderr)
        return 1

    if load_dotenv is not None:
        load_dotenv()

    access_token = os.environ.get("META_ACCESS_TOKEN")
    api_version = os.environ.get("META_API_VERSION", "v21.0")
    if not access_token:
        print("META_ACCESS_TOKEN is not set.", file=sys.stderr)
        return 1

    accounts_by_key = _discover_accounts()
    if not accounts_by_key:
        print("No Meta ad accounts configured.", file=sys.stderr)
        return 1
    account_key = args.account or sorted(accounts_by_key, key=int)[0]
    if account_key not in accounts_by_key:
        print(f"No account with key '{account_key}'.", file=sys.stderr)
        return 1
    account = accounts_by_key[account_key]

    registry_fields = _load_registry_field_universe(META_REGISTRY_PATH)
    doc_fields = _load_doc_field_universe(FIELDS_XLSX_PATH)
    universe = sorted(registry_fields | doc_fields)

    total_calls = len(universe) * len(levels)
    print(f"Account: {account.key} ({account.name})")
    print(f"Field universe: {len(universe)} ({len(registry_fields)} in registry, {len(doc_fields)} in doc, "
          f"{len(registry_fields & doc_fields)} overlap)")
    print(f"Levels: {levels} -> {total_calls} total (field, level) requests")
    print(f"Testing individually (concurrency={args.concurrency}), date_preset={args.date_preset}\n")

    base_url = f"https://graph.facebook.com/{api_version}"
    t0 = time.monotonic()
    results = asyncio.run(
        _run(universe, levels, base_url=base_url, account=account, access_token=access_token,
             date_preset=args.date_preset, concurrency=args.concurrency)
    )
    elapsed = time.monotonic() - t0

    results.sort(key=lambda r: (r["field"], r["level"]))
    passed = [r for r in results if r["status"] == "passed"]
    errored = [r for r in results if r["status"] != "passed"]

    print(f"\nDone in {elapsed:.1f}s. Passed: {len(passed)}  Errored: {len(errored)}\n")

    def status_string(r: dict[str, Any]) -> str:
        if r["status"] == "passed":
            return "passed"
        code = r.get("error_code", "")
        msg = r.get("error_message", "")
        return f"error {code}: {msg}" if code else f"error: {msg}"

    # {field: {"account_level": "passed"|"error: ...", "campaign_level": ..., ...}}
    # -- the exact requested shape.
    nested_map: dict[str, dict[str, str]] = {}
    for r in results:
        nested_map.setdefault(r["field"], {})[f"{r['level']}_level"] = status_string(r)

    OUT_JSON_PATH.parent.mkdir(exist_ok=True)
    with open(OUT_JSON_PATH, "w") as f:
        json.dump(nested_map, f, indent=2, sort_keys=True)

    with open(OUT_JSONL_PATH, "w") as f:
        for r in results:
            in_registry = r["field"] in registry_fields
            f.write(json.dumps({**r, "in_current_registry": in_registry}) + "\n")

    with open(OUT_CSV_PATH, "w", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=CSV_COLUMNS)
        writer.writeheader()
        for r in results:
            writer.writerow({
                "field": r["field"],
                "level": r["level"],
                "status": r["status"],
                "in_current_registry": r["field"] in registry_fields,
                "error_code": r.get("error_code", ""),
                "error_subcode": r.get("error_subcode", ""),
                "error_type": r.get("error_type", ""),
                "error_message": r.get("error_message", ""),
            })

    print(f"Written: {OUT_JSON_PATH.relative_to(REPO_ROOT)}, {OUT_CSV_PATH.relative_to(REPO_ROOT)}, "
          f"{OUT_JSONL_PATH.relative_to(REPO_ROOT)}")

    # Fields whose pass/fail differs across levels -- the interesting
    # cross-level cases (level-restricted metrics, either already
    # documented in LEVEL_RESTRICTED_METRICS or newly discovered here).
    mixed = {
        field: levels_map
        for field, levels_map in nested_map.items()
        if len({v == "passed" for v in levels_map.values()}) > 1
    }
    if mixed:
        print(f"\n{len(mixed)} field(s) pass at SOME levels but not others:")
        for field, levels_map in sorted(mixed.items()):
            summary = ", ".join(f"{lvl}={'OK' if v == 'passed' else 'FAIL'}" for lvl, v in sorted(levels_map.items()))
            print(f"  {field}: {summary}")

    if errored:
        print(f"\n{len(errored)} (field, level) error(s) total -- see {OUT_CSV_PATH.relative_to(REPO_ROOT)} for full detail")

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
